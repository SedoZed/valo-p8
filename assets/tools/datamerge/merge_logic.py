import logging
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# ---------------------------------------------------------
# Charger Excel en éclatant les cellules fusionnées
# ---------------------------------------------------------
def load_excel_clean(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    # Grille brute
    grid = [[None for _ in range(max_col)] for _ in range(max_row)]

    # Remplir la grille
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            grid[r-1][c-1] = ws.cell(row=r, column=c).value

    # Défusionner
    for merged in ws.merged_cells.ranges:
        value = ws.cell(row=merged.min_row, column=merged.min_col).value
        for r in range(merged.min_row, merged.max_row + 1):
            for c in range(merged.min_col, merged.max_col + 1):
                grid[r-1][c-1] = value

    df = pd.DataFrame(grid)

    # Supprimer lignes et colonnes vides
    df = df.dropna(how="all")
    df = df.dropna(axis=1, how="all")

    # ✅ Renommer en garantissant l'unicité
    new_cols = []
    seen = {}

    for i in range(df.shape[1]):
        base = f"col_{i}"
        name = base

        if name in seen:
            seen[name] += 1
            name = f"{base}_{seen[name]}"
        else:
            seen[name] = 0

        new_cols.append(name)

    df.columns = new_cols

    return df

# ---------------------------------------------------------
# Fusion
# ---------------------------------------------------------
def merge_files(
    file1_path,
    file2_path,
    key_column_file1,
    key_column_file2,
    source_column,
    target_column,
    output_path
):
    logger.info("=== DÉBUT MERGE ===")

    df1 = load_excel_clean(file1_path)
    df2 = load_excel_clean(file2_path)

    logger.info(f"Colonnes fichier 1 : {df1.columns.tolist()}")
    logger.info(f"Colonnes fichier 2 : {df2.columns.tolist()}")

    # ✅ DEBUG : afficher les choix utilisateur
    print("\n=== DEBUG CHOIX UTILISATEUR ===")
    print("KEY FILE 1 =", key_column_file1)
    print("KEY FILE 2 =", key_column_file2)
    print("SOURCE =", source_column)
    print("TARGET =", target_column)

    # ✅ Empêcher les doublons de sélection
    if key_column_file1 == source_column:
        raise ValueError("Erreur : la clé du fichier 1 et la colonne source sont identiques.")

    if key_column_file2 == target_column:
        raise ValueError("Erreur : la clé du fichier 2 et la colonne destination sont identiques.")

    # Vérification existence colonnes
    for col, name, df in [
        (key_column_file1, "clé fichier 1", df1),
        (key_column_file2, "clé fichier 2", df2),
        (source_column, "source fichier 1", df1),
    ]:
        if col not in df.columns:
            raise ValueError(f"Colonne '{col}' introuvable dans {name}")

    # Normalisation des clés
    df1[key_column_file1] = df1[key_column_file1].astype(str).str.strip().str.upper()
    df2[key_column_file2] = df2[key_column_file2].astype(str).str.strip().str.upper()

    # ✅ DEBUG : afficher colonnes après normalisation
    print("\n=== DEBUG DF1 ===")
    print(df1.columns.tolist())
    print(df1.head())

    print("\n=== DEBUG DF2 ===")
    print(df2.columns.tolist())
    print(df2.head())

    # ✅ On NE renomme plus rien → on fusionne avec left_on / right_on
    left_key = key_column_file2
    right_key = key_column_file1

    print("\n=== DEBUG CLÉS DE FUSION ===")
    print("LEFT KEY =", left_key)
    print("RIGHT KEY =", right_key)

    # ✅ Regrouper les doublons dans df1
    df1_grouped = df1.groupby(right_key)[source_column].apply(
        lambda x: " | ".join(sorted(set(str(v) for v in x if pd.notna(v))))
    ).reset_index()

    # Renommer la colonne source pour la fusion
    df1_grouped = df1_grouped.rename(columns={source_column: "source_value"})

    print("\n=== DEBUG DF1 GROUPED ===")
    print(df1_grouped.head())

    # ✅ Fusion
    df_merged = df2.merge(
        df1_grouped,
        left_on=left_key,
        right_on=right_key,
        how="left"
    )

    # ✅ Création colonne destination si absente
    if target_column not in df_merged.columns:
        df_merged[target_column] = pd.NA

    # ✅ Remplissage à partir de 'source_value'
    filled_count = 0
    for idx, row in df_merged.iterrows():
        if pd.isna(row[target_column]) or str(row[target_column]).strip() == "":
            if "source_value" in row.index and not pd.isna(row["source_value"]):
                df_merged.at[idx, target_column] = row["source_value"]
                filled_count += 1

    logger.info(f"Lignes remplies : {filled_count}")

    # ✅ Sauvegarde
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_merged.to_excel(output_path, index=False)

    logger.info("=== FIN MERGE ===")
